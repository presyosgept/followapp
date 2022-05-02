
from django.http import HttpResponse
from json import encoder
from typing import Counter
from django.db.models.fields import TimeField
from django.http import HttpResponse, request
from django.http.response import Http404, HttpResponsePermanentRedirect, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Max, Min
from django.contrib.auth import authenticate, login, logout

from django.contrib import messages

from django.contrib.auth.decorators import login_required

from .utilities import create_notification, create_feedback
from django.views import generic
from django.utils import timezone
from datetime import date, datetime, timedelta

import datetime as dt

from .forms import SetScheduleCounselorForm, FilterForm, DeleteDepartmentForm, AddDepartmentForm, DeleteSchoolOfficeForm, AddSchoolOfficeForm, ProgramForm, CalendarForm, StudentInfoForm, OfferingForm, StudentSetSchedForm, CounselorFeedbackForm, VerificationForm, AccountCreatedForm, AccountsForm, CounselorForm, TeachersReferralForm, StudentsForm, CreateUserForm, SubjectOfferedForm, FacultyloadForm, StudentsloadForm
from .models import NewTime, SetScheduleCounselor, NewDepartment, Calendar, StudentInfo, Offering, StudentSetSched, NotificationFeedback, CounselorFeedback, SubjectWithSem, Semester, AccountCreated, Faculty, Counselor, Notification, Counselor, TeachersReferral,  SubjectOffered, Facultyload, Studentsload

from .resources import TimeResource, NewDepartmentResource, SubjectWithSemResource, SemesterResource, StudentsloadResource, FacultyResource, CounselorResource, TeachersReferralResource, SubjectOfferedResource, FacultyloadResource
from tablib import Dataset

# Create your views here.

from django.shortcuts import render
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from . serializers import FacultySerializers, ResultSerializer, Result, UserSerializer

from .models import AccountsApi, Subject, OfferCode, SchoolOffices, DegreeProgram, AllStudent

from .resources import SubjectResource, OfferCodeResource, SchoolOfficesResource, DegreeProgramResource, AllStudentResource
import openpyxl


from django.views.generic import View
from django.shortcuts import redirect
from django.contrib import messages
from django.core.mail import send_mail, send_mass_mail
from django.conf import settings
from django.core.mail import get_connection
from django.core.mail.message import EmailMessage
import random

from .serializers import UserSerializer, ActorSerializer, Actor
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAdminUser
from django.contrib.auth.models import User
from django.shortcuts import render

from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser
from rest_framework import status
from rest_framework import generics
from rest_framework.decorators import api_view


# Create your views here.

# class studentsList(APIView):
#     def get(self, request):
#         stud = Students.objects.all()
#         serializer=studentsSerializers(stud, many=True)
#         return Response({'students':serializer.data})

class CounselorList(APIView):
    def get(self, request):
        # couns = Faculty.objects.all()
        # serializer=FacultySerializers(couns, many=True)
        obj = Result(bool1=True)
        serializer = ResultSerializer(obj)
        return Response(serializer.data)

# api


class RegisterApi(APIView):
    def get(self, request, id, email):
        char = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
        for x in range(0, 1):
            code = ''
            for x in range(0, 8):
                code_char = random.choice(char)
                code = code + code_char

        userTeacher = Faculty.objects.filter(employee_id=id).first()
        userStudent = AllStudent.objects.filter(studnumber=id).first()
        exist = AccountsApi.objects.filter(id_number=id).first()
        flag = 0
        if exist is not None:
            flag = 1
        obj = Result(bool1=False)
        if userTeacher is not None or userStudent is not None:
            if flag == 0:
                connection = get_connection(use_tls=True,
                                            host='smtp.gmail.com',
                                            port=587,
                                            username='followapp2021@gmail.com',
                                            password='hellocapstone2')
                EmailMessage(
                    "Verification Code",
                    "This is your verification code: " + code,
                    'followapp2021@gmail.com',
                    [
                        email,
                    ], connection=connection).send()
                value = AccountsApi(id_number=id, email=email, code=code)
                value.save()
                obj = Result(bool1=True)
                serializer = ResultSerializer(obj)
                return JsonResponse({"result": "Account Created"}, status=status.HTTP_200_OK)
            else:
                return JsonResponse({"result": "Account is Existing"}, status=status.HTTP_400_BAD_REQUEST)

        serializer = ResultSerializer(obj)
        return JsonResponse({"result": "Invalid Account"}, status=status.HTTP_400_BAD_REQUEST)


class VerificationApi(APIView):
    def get(self, request, id, code):
        user = AccountsApi.objects.filter(id_number=id, code=code).first()
        obj = Result(bool1=False)
        if user is not None:
            obj = Result(bool1=True)
            serializer = ResultSerializer(obj)
            return JsonResponse(serializer.data, status=status.HTTP_200_OK)
        else:
            serializer = ResultSerializer(obj)
            return JsonResponse(serializer.data, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'POST', 'DELETE'])
def signup_api(request):
    if request.method == 'POST':
        user_data = JSONParser().parse(request)
        user_serializer = UserSerializer(data=user_data)

    username = str(user_data["username"])
    user = AccountsApi.objects.filter(id_number=username).first()
    if user_serializer.is_valid():
        if user is not None:
            user_serializer.save()
            return JsonResponse(user_serializer.data, status=status.HTTP_200_OK)
        else:
            return JsonResponse({"username": "Username Not Valid"}, status=status.HTTP_400_BAD_REQUEST)
    return JsonResponse(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'POST', 'DELETE'])
def login_api(request):
    if request.method == 'POST':
        user_data = JSONParser().parse(request)

    username = str(user_data["username"])
    password = str(user_data["password"])
    user = User.objects.filter(username=username, password=password).first()
    userTeacher = Faculty.objects.filter(employee_id=username).first()
    userStudent = AllStudent.objects.filter(studnumber=username).first()

    if user is not None:
        if userTeacher is not None:
            obj = Actor(actor="teacher")
            serializer = ActorSerializer(obj)
            return JsonResponse(serializer.data, status=status.HTTP_200_OK)
        if userStudent is not None:
            obj = Actor(actor="learner")
            serializer = ActorSerializer(obj)
            return JsonResponse(serializer.data, status=status.HTTP_200_OK)
    else:
        return JsonResponse({'actor': "no account"}, status=status.HTTP_400_BAD_REQUEST)


# couns = Faculty.objects.all()
    # serializer=FacultySerializers(couns, many=True)
# ({'students':serializer.data})

# from django.contrib.auth.forms import User
# class RegisterApi(APIView):
#     def get(request, id, password):
#         user = User.objects.create(
#                 username=id,
#                 password = password
#             )
#         user.save()
#         return user
#         # obj = Result(bool1 = False)
#         # accs = AccountsApi.objects.all()
#         # for check in accs:
#         #     if(check.id_number == id):
#         #         user = authenticate(request, username=id, password=password)
#         #         obj = Result(bool1 = True)

#         # serializer = ResultSerializer(obj)
#         # return Response(serializer.data)

#     # def create(self, validated_data):

# api

# global variables
counselorNotif = 0
studentNotif = 0
teacherNotif = 0
count = 0
count1 = 0
formm = AccountsForm()
feedback_id = 0
global sem
global sy
global dep
global school


def register(request):
    form = AccountCreatedForm()
    if request.method == 'POST':
        char = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
        for x in range(0, 1):
            code = ''
            for x in range(0, 8):
                code_char = random.choice(char)
                code = code + code_char
        username = request.POST.get('id_number')
        email = request.POST.get('email')
        qs_faculty = Faculty.objects.all()
        qs_student = AllStudent.objects.all()
        qs_acc = AccountCreated.objects.all()

        if username == 'followapp':
            connection = get_connection(use_tls=True,
                                        host='smtp.gmail.com',
                                        port=587,
                                        username='followapp2021@gmail.com',
                                        password='hellocapstone2')
            EmailMessage(
                "Verification Code",
                "This is your verification code: " + code,
                'followapp2021@gmail.com',
                [
                    email,
                ], connection=connection).send()
            value = AccountCreated(
                id_number=username, email=email, password=code)
            value.save()
            return redirect('verification_code')

        exist = 0
        for acc in qs_acc:
            if username == acc.id_number:
                exist = 1

        if exist == 0:
            flag = 0
            for user in qs_student:
                if username == user.studnumber:
                    flag = 1
            for user in qs_faculty:
                if username == user.employee_id:
                    flag = 1
            if flag == 1:
                connection = get_connection(use_tls=True,
                                            host='smtp.gmail.com',
                                            port=587,
                                            username='followapp2021@gmail.com',
                                            password='hellocapstone2')
                EmailMessage(
                    "Verification Code",
                    "This is your verification code: " + code,
                    'followapp2021@gmail.com',
                    [
                        email,
                    ], connection=connection).send()
                value = AccountCreated(
                    id_number=username, email=email, password=code)
                value.save()
                messages.info(request, "Check Gmail for Code")
                return redirect('verification_code')
            else:
                messages.info(request, "Account Not Valid")
        else:
            messages.info(request, "Account Already Existed")
    else:
        AccountsForm()

    return render(request, 'register.html', {'form': form})


def verification_code(request):
    form = VerificationForm()
    code = request.POST.get('code')
    if request.method == 'POST':
        qs_account = AccountCreated.objects.all()
        flag = 0
        for vcode in qs_account:
            if (code == vcode.password):
                flag = 1

        if(flag == 1):
            return redirect('signup')
        else:
            VerificationForm()
            messages.info(request, 'Invalid Code')
    else:
        VerificationForm()

    return render(request, 'verification.html', {'form': form})


def signup(request):
    if request.user.is_authenticated:
        return redirect('login')
    else:
        form = CreateUserForm()
        if request.method == 'POST':
            username = request.POST.get('username')
            qs_account = AccountCreated.objects.all()
            exist = 0
            for user in qs_account:
                if user.id_number == username:
                    exist = 1
            if (exist == 1):
                flag = 0
                qs = AllStudent.objects.all()
                for student in qs:
                    if student.studnumber == username:
                        flag = 1
                        stud = AllStudent.objects.get(studnumber=username)
                        studentInfo = StudentInfo(firstname=stud.firstname,
                                                  lastname=stud.lastname, middlename=stud.middlename, studnumber=stud.studnumber,
                                                  degree_program=stud.degree_program, year=stud.year, student_email=stud.student_email)
                        studentInfo.save()
                if flag == 1:
                    form = CreateUserForm(request.POST)
                    if form.is_valid():
                        form.save()
                        user = form.cleaned_data.get('username')
                        messages.info(
                            request, 'Student Account was created for ' + user)
                        return redirect('login')
                elif flag == 0:
                    qs = Faculty.objects.all()
                    for teacher in qs:
                        if teacher.employee_id == username:
                            form = CreateUserForm(request.POST)
                            if form.is_valid():
                                form.save()
                                user = form.cleaned_data.get('username')
                                messages.info(
                                    request, 'Account was created for ' + user)
                                return redirect('login')

                if username == 'followapp':
                    form = CreateUserForm(request.POST)
                    if form.is_valid():
                        form.save()
                        user = form.cleaned_data.get('username')
                        messages.info(
                            request, 'Admin Account was created for ' + user)
                        return redirect('login')

                messages.info(request, 'Check Credentials Account Not Created')

            else:
                messages.info(request, 'Check Credentials Account Not Created')

    return render(request, 'signup.html', {'form': form})


def loginPage(request):
    if request.user.is_authenticated:
        return redirect('login')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')

            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                if request.method == 'POST':
                    flag = 0
                    username = request.POST.get('username')
                    qs = AllStudent.objects.all()
                    for student in qs:
                        if student.studnumber == username:
                            if student.role == 'learner':
                                flag = 1
                    if flag == 1:
                        request.session['username'] = username
                        # return redirect('student_home_view')
                        check = StudentInfo.objects.all()
                        if check is not None:
                            stud = StudentInfo.objects.get(studnumber=username)
                            if stud.status == 'undone':
                                return redirect('student_add_info')
                            else:
                                return redirect('student_home_view')
                        else:
                            return redirect('student_add_info')

                    else:
                        qs = Faculty.objects.all()
                        for teacher in qs:
                            if teacher.employee_id == username:
                                if teacher.role == 'teacher':
                                    request.session['username'] = username
                                    flag = 2
                                elif teacher.role == 'counselor':
                                    request.session['username'] = username
                                    flag = 3
                                elif teacher.role == 'director':
                                    request.session['username'] = username
                                    flag = 4
                    if flag == 2:
                        return redirect('teacher_home_view')
                    if flag == 3:
                        return redirect('counselor_home_view')
                    if flag == 4:
                        return redirect('director_home_view')
                    if username == 'followapp':
                        return redirect('admin_home_view')
                    if username == 'upload':
                        return redirect('uploaddb_home_view')

            else:
                messages.info(request, 'Username or Password is Incorrect')

        return render(request, 'login.html', {})


def logoutUser(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def home(request, *args, **kwargs):
    return render(request, "welcomepage.html", {})


# director


@login_required(login_url='login')
def director_home_view(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    return render(request, "director/director_home.html", {"object": director_name})


@login_required(login_url='login')
def director_counselor(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    qs = Faculty.objects.filter(role='counselor')
    return render(request, "director/director_counselor.html", {"object_list": qs, "object": director_name})


@login_required(login_url='login')
def view_stat_by_counselor(request, id):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    stat_details = TeachersReferral.objects.filter(counselor=id)
    stat = stat_details.count()
    return render(request, "director/view_stat_by_counselor.html", {"object": director_name, 'stat_details': stat_details, 'stat': stat})


@login_required(login_url='login')
def director_degree_program(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    degree = DegreeProgram.objects.all()
    return render(request, "director/director_degree_program.html", {"object": director_name, 'degree': degree})


@login_required(login_url='login')
def view_stat_by_degree_program(request, degree):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    stat_details = TeachersReferral.objects.filter(degree_program=degree)
    stat = stat_details.count()
    return render(request, "director/view_stat_by_degree_program.html", {"object": director_name, 'stat_details': stat_details, 'stat': stat})


@login_required(login_url='login')
def director_offering(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    global sem
    global sy
    global dep
    form = OfferingForm()
    if request.method == "POST":
        form = OfferingForm(request.POST)
        if form.is_valid():
            form.save()
            semester = form['semester'].value()
            schoolyear = form['school_year'].value()
            depa_choice = form['depa_choice'].value()
            dep = depa_choice
            sem = semester
            sy = schoolyear
            return redirect(director_view_offering)

    return render(request, "director/offering.html", {"object": director_name, "form": form})


@login_required(login_url='login')
def director_view_offering(request):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    global sem
    global sy
    global dep
    semester = ''
    if sem == "1ST SEM":
        semester = "101"
    elif sem == "2ND SEM":
        semester = "201"

    depa = NewDepartment.objects.get(department_name=dep)
    allsubj = Subject.objects.filter(department_id_id=depa.department_id)
    qs = OfferCode.objects.filter(sem_id=semester, academic_year=sy)
    offering = []
    for a in allsubj:
        for b in qs:
            if(a.subject_code == b.subject_code):
                concatenatedStringofDays = ""
                for object1 in b.days:
                    concatenatedStringofDays += object1+" "
                offering.append(OfferCode(offer_code=b.offer_code, days=concatenatedStringofDays, start_time=b.start_time, end_time=b.end_time, room=b.room,
                                          subject_code=b.subject_code, sem_id=b.sem_id, academic_year=b.academic_year))

    return render(request, "director/viewoffering.html", {"object": director_name, "forms": offering})


@login_required(login_url='login')
def view_stats(request, *args, **kwargs):
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    couns = Counselor.objects.all()
    today = date.today()
    day = today.strftime("%Y-%m-%d")
    year = today.strftime("%Y")
    month = today.strftime("%Y-%m")
    textMon = today.strftime('%B')
    textDay = today.strftime("%d")
    referals = TeachersReferral.objects.all()
    stat_month = 0
    stat_day = 0
    stat_year = 0
    for stud in referals:
        getmonth = stud.date.strftime("%Y-%m")
        getday = stud.date.strftime("%Y-%m-%d")
        getyear = stud.date.strftime("%Y")
        if month == getmonth:
            stat_month += 1
        if day == getday:
            stat_day += 1
        if year == getyear:
            stat_year += 1

    return render(request, "director/view_stats.html", {"textYear": year, "textDay": textDay, "textMonth": textMon, "offer": offer, "couns": couns, "object": director_name, "stat_month": stat_month, "stat_day": stat_day, "stat_year": stat_year})


@login_required(login_url='login')
def view_another_stats(request, *args, **kwargs):
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    newDate = Calendar.objects.last()
    today = newDate.pickedDate
    day = today.strftime("%Y-%m-%d")
    year = today.strftime("%Y")
    month = today.strftime("%Y-%m")
    textMon = today.strftime('%B')
    textDay = today.strftime("%d")
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    couns = Counselor.objects.all()
    referals = TeachersReferral.objects.all()
    stat_month = 0
    stat_day = 0
    stat_year = 0
    for stud in referals:
        getmonth = stud.date.strftime("%Y-%m")
        getday = stud.date.strftime("%Y-%m-%d")
        getyear = stud.date.strftime("%Y")
        if month == getmonth:
            stat_month += 1
        if day == getday:
            stat_day += 1
        if year == getyear:
            stat_year += 1
    return render(request, "director/view_another_stats.html", {"textYear": year, "textDay": textDay, "textMonth": textMon, "offer": offer, "couns": couns, "object": director_name, "stat_month": stat_month, "stat_day": stat_day, "stat_year": stat_year})


@login_required(login_url='login')
def view_another_stat_specific_counselor(request, id):
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    newDate = Calendar.objects.last()
    today = newDate.pickedDate
    day = today.strftime("%Y-%m-%d")
    year = today.strftime("%Y")
    month = today.strftime("%Y-%m")
    textMon = today.strftime('%B')
    textDay = today.strftime("%d")
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    couns = Counselor.objects.get(employee_id=id)
    referals = TeachersReferral.objects.filter(counselor=id, date=today)
    stat_month = 0
    stat_day = 0
    stat_year = 0
    for stud in referals:
        getmonth = stud.date.strftime("%Y-%m")
        getday = stud.date.strftime("%Y-%m-%d")
        getyear = stud.date.strftime("%Y")
        if month == getmonth:
            stat_month += 1
        if day == getday:
            stat_day += 1
        if year == getyear:
            stat_year += 1
    return render(request, "director/view_stat_specific_counselor.html", {"textYear": year, "textDay": textDay, "textMonth": textMon, "offer": offer, "id": id, "couns": couns, "referals": referals, "object": director_name, "stat_month": stat_month, "stat_day": stat_day, "stat_year": stat_year})


@login_required(login_url='login')
def view_stat_specific_counselor(request, id):
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    couns = Counselor.objects.get(employee_id=id)
    today = date.today()
    day = today.strftime("%Y-%m-%d")
    year = today.strftime("%Y")
    textMon = today.strftime('%B')
    textDay = today.strftime("%d")
    month = today.strftime("%Y-%m")
    referals = TeachersReferral.objects.filter(counselor=id)
    stat_month = 0
    stat_day = 0
    stat_year = 0
    for stud in referals:
        getmonth = stud.date.strftime("%Y-%m")
        getday = stud.date.strftime("%Y-%m-%d")
        getyear = stud.date.strftime("%Y")
        if month == getmonth:
            stat_month += 1
        if day == getday:
            stat_day += 1
        if year == getyear:
            stat_year += 1
    return render(request, "director/view_another_stat_specific_counselor.html", {"textYear": year, "textDay": textDay, "textMonth": textMon, "offer": offer, "id": id, "couns": couns, "referals": referals, "object": director_name, "stat_month": stat_month, "stat_day": stat_day, "stat_year": stat_year})


@login_required(login_url='login')
def director_assign_counselor(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    qs = Faculty.objects.filter(role='counselor')
    return render(request, "director/assign_counselor.html", {"object_list": qs, "object": director_name})


global director_pk


@login_required(login_url='login')
def director_fillinForm(request, pk):
    global director_pk
    global school
    director_pk = pk
    user = request.session.get('username')
    counselor = Counselor.objects.get(employee_id=pk)
    form = CounselorForm(instance=counselor)
    director_name = Faculty.objects.get(employee_id=user)
    program_not_to_assign = []
    check = []
    flag = 0
    check = 0

    if request.method == "POST":
        form = CounselorForm(request.POST, instance=counselor)
        if form.is_valid():
            counselorCheck = Counselor.objects.all()
            program = form.cleaned_data['program_designation']
            for object in counselorCheck:
                if object.program_designation is not None:
                    check = 1
            if check == 1:
                for object in counselorCheck:
                    if(object.employee_id != pk):
                        for object1 in object.program_designation:
                            if(object1 in program):
                                program_not_to_assign.append(object1)
                                flag = 1
            if(flag == 0):
                form.save()
                program_not_to_assign = []
                messages.info(
                    request, 'Successfully Assigned The Counselor')
            else:
                concatenatedStringofPrograms = ""
                for object in program_not_to_assign:
                    concatenatedStringofPrograms += object+" "
                messages.info(
                    request, 'Program already taken: ' + str(concatenatedStringofPrograms))
                program_not_to_assign = []
                flag = 0
    return render(request, "director/form.html", {"form1": form, "object": director_name})


@login_required(login_url='login')
def director_choose_program(request):
    global school
    global director_pk
    programs = []
    progform = ProgramForm()
    if request.method == "POST":
        progform = ProgramForm(request.POST)
        designation = progform['program'].value()
        if progform.is_valid():
            designation = progform['program'].value()
            progform.save()

    degreeprogram = DegreeProgram.objects.all()
    for degree in degreeprogram:
        if degree.school_code == school:
            t = Counselor.objects.get(employee_id=director_pk)
            t.program_designation = degree.program_code
            t.save()
            programs.append(DegreeProgram(
                program_code=degree.program_code, program_name=degree.program_name))
    user = request.session.get('username')
    director_name = Faculty.objects.get(employee_id=user)
    qs = Faculty.objects.filter(role='counselor')
    return render(request, "director/director_choose_program.html", {"forms": programs, "progform": progform, "object": director_pk})

# director

# admin


@login_required(login_url='login')
def admin_home_view(request, *args, **kwargs):
    return render(request, "admin/admin_home.html", {})


@login_required(login_url='login')
def view_enrolled_students_via_degree(request):
    degree = DegreeProgram.objects.all()
    return render(request, "admin/view_enrolled_students_via_degree.html", {"degree": degree})


@login_required(login_url='login')
def view_enrolled_students(request, id):
    degree = DegreeProgram.objects.get(program_id=id)
    student = AllStudent.objects.filter(degree_program_id=id)
    return render(request, "admin/view_enrolled_students.html", {"student": student, "degree": degree})


@login_required(login_url='login')
def view_schooloffices(request):
    school = SchoolOffices.objects.all()
    addschoolForm = AddSchoolOfficeForm()
    deleteschoolForm = DeleteSchoolOfficeForm()
    flag = 0
    if request.method == "POST":
        addschoolForm = AddSchoolOfficeForm(request.POST)
        deleteschoolForm = DeleteSchoolOfficeForm(request.POST)
        if addschoolForm.is_valid():
            code = addschoolForm.cleaned_data['school_code']
            name = addschoolForm.cleaned_data['school_office_name']
            newdata = SchoolOffices(
                school_code=code.upper(), school_office_name=name)
            newdata.save()
            messages.info(request, 'Successfully Added')

        if deleteschoolForm.is_valid():
            delete = deleteschoolForm['schoolform_code'].value()
            schoolOffc = SchoolOffices.objects.all()
            for object in schoolOffc:
                if object.school_code == delete.upper():
                    flag = 1
                    object.delete()
                    messages.info(request, 'Successfully Deleted')
            if(flag == 0):
                messages.info(request, 'No Code Found')
        addschoolForm = AddSchoolOfficeForm()
        deleteschoolForm = DeleteSchoolOfficeForm()

    return render(request, "admin/view_schooloffices.html", {"school": school, "addschoolForm": addschoolForm, "deleteschoolForm": deleteschoolForm})


@login_required(login_url='login')
def view_department(request):
    school = SchoolOffices.objects.all()
    return render(request, "admin/view_department.html", {"school": school})


@login_required(login_url='login')
def add_department(request, school):
    depa = NewDepartment.objects.filter(school_code_id=school)
    addDepaForm = AddDepartmentForm()
    deleteDepaForm = DeleteDepartmentForm()
    flag = 0
    exist = 0
    if request.method == "POST":
        addDepaForm = AddDepartmentForm(request.POST)
        deleteDepaForm = DeleteDepartmentForm(request.POST)
        if addDepaForm.is_valid():
            count = NewDepartment.objects.count()
            department_name_form1 = addDepaForm.cleaned_data['department_name_form']
            checking = NewDepartment.objects.all()
            for obj in checking:
                if obj.department_name == department_name_form1:
                    exist = 1
            if exist == 0:
                t = NewDepartment(department_id=count+1, department_name=department_name_form1,
                                  school_code_id=school)
                t.save()
                addDepaForm = AddDepartmentForm()
                deleteDepaForm = DeleteDepartmentForm()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Department Name is Existing')

        if deleteDepaForm.is_valid():
            delete = deleteDepaForm.cleaned_data['del_department_name_form']
            deleteDepa = NewDepartment.objects.filter(school_code_id=school)
            idOfIdDeleted = NewDepartment.objects.get(
                department_name=delete)
            id = idOfIdDeleted.department_id
            for object in deleteDepa:
                if object.department_name == delete:
                    flag = 1
                    object.delete()
                    departmentList = NewDepartment.objects.all()
                    for object in departmentList:
                        if int(object.department_id) > int(id):
                            ty = NewDepartment.objects.get(
                                department_id=object.department_id)
                            newId = int(ty.department_id) - 1
                            ty.department_id = str(newId)
                            ty.save()
                    departmentList = NewDepartment.objects.all()
                    store = []
                    for obj in departmentList:
                        store.append(NewDepartment(
                            department_id=obj.department_id))
                    max = 0
                    for a in store:
                        if int(a.department_id) > max:
                            max = int(a.department_id)

                    addDepaForm = AddDepartmentForm()
                    deleteDepaForm = DeleteDepartmentForm()
                    new = NewDepartment.objects.get(department_id=str(max))
                    new.delete()
                    departmentList = NewDepartment.objects.all()
                    messages.info(request, 'Successfully Deleted')

            if(flag == 0):
                messages.info(request, 'Not Existing')

    department = NewDepartment.objects.filter(school_code_id=school)
    return render(request, "admin/add_department.html", {"depa": department, "addDepaForm": addDepaForm, "deleteDepaForm": deleteDepaForm, "school": school})


@login_required(login_url='login')
def admin_offering(request, *args, **kwargs):
    global sem
    global sy
    global dep
    form = OfferingForm()
    if request.method == "POST":
        form = OfferingForm(request.POST)
        if form.is_valid():
            form.save()
            semester = form['semester'].value()
            schoolyear = form['school_year'].value()
            depa_choice = form['depa_choice'].value()
            dep = depa_choice
            sem = semester
            sy = schoolyear
            return redirect(admin_view_offering)

    return render(request, "admin/offering.html", {"form": form})


@login_required(login_url='login')
def admin_view_offering(request):
    global sem
    global sy
    global dep
    semester = ''
    if sem == "1ST SEM":
        semester = "101"
    elif sem == "2ND SEM":
        semester = "201"

    depa = NewDepartment.objects.get(department_name=dep)
    allsubj = Subject.objects.filter(department_id_id=depa.department_id)
    qs = OfferCode.objects.filter(sem_id=semester, academic_year=sy)
    offering = []
    for a in allsubj:
        for b in qs:
            if(a.subject_code == b.subject_code):
                concatenatedStringofDays = ""
                for object1 in b.days:
                    concatenatedStringofDays += object1+" "
                offering.append(OfferCode(offer_code=b.offer_code, days=concatenatedStringofDays, start_time=b.start_time, end_time=b.end_time, room=b.room,
                                          subject_code=b.subject_code, sem_id=b.sem_id, academic_year=b.academic_year))

    return render(request, "admin/viewoffering.html", {"forms": offering})


@login_required(login_url='login')
def upload_studentsload(request):
    try:
        if request.method == 'POST':
            StudentsloadResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 3):
                for data in imported_data:
                    value = Studentsload(
                        data[0],
                        data[1],
                        data[2],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_studentsload.html")


@login_required(login_url='login')
def upload_students(request):
    try:
        if request.method == 'POST':
            AllStudentResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']
            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row
            if(col == 8):
                for data in imported_data:
                    value = AllStudent(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                        data[6],
                        data[7],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_students.html")


@login_required(login_url='login')
def upload_faculty(request):
    try:
        if request.method == 'POST':
            FacultyResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 6):
                for data in imported_data:
                    value = Faculty(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_faculty.html")


@login_required(login_url='login')
def upload_facultyload(request):
    try:
        if request.method == 'POST':
            FacultyloadResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row
            if(col == 3):
                for data in imported_data:
                    value = Facultyload(
                        data[0],
                        data[1],
                        data[2],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_facultyload.html")


@login_required(login_url='login')
def uploaddb_counselor(request):
    try:
        if request.method == 'POST':
            CounselorResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row
            if(col == 5):
                for data in imported_data:
                    value = Counselor(
                        data[0],
                        data[1],
                        data[2]
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_counselor.html")


@login_required(login_url='login')
def offercode(request):
    school = SchoolOffices.objects.all()
    return render(request, "admin/offercode.html", {"school": school})


@login_required(login_url='login')
def uploaddb_offercode(request, school_office):
    try:
        if request.method == 'POST':
            OfferCodeResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 8):
                for data in imported_data:
                    value = OfferCode(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                        data[6],
                        data[7],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_offercode.html", {"school_office": school_office})


@login_required(login_url='login')
def uploaddb_semester(request):
    try:
        if request.method == 'POST':
            SemesterResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']
            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 2):
                for data in imported_data:
                    value = Semester(
                        data[0],
                        data[1],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_semester.html")


@login_required(login_url='login')
def uploaddb_allsubject(request):
    try:
        if request.method == 'POST':
            SubjectResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 4):
                for data in imported_data:
                    value = Subject(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_allsubject.html")


@login_required(login_url='login')
def uploaddb_degreeprogram(request):
    try:
        if request.method == 'POST':
            DegreeProgramResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 4):
                for data in imported_data:
                    value = DegreeProgram(
                        data[0],
                        data[1],
                        data[2],
                        data[3]
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_degreeprogram.html")


@login_required(login_url='login')
def uploaddb_department(request):
    try:
        if request.method == 'POST':
            NewDepartmentResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 3):
                for data in imported_data:
                    value = NewDepartment(
                        data[0],
                        data[1],
                        data[2]
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_department.html")


@login_required(login_url='login')
def uploaddb_schooloffices(request):
    try:
        if request.method == 'POST':
            SchoolOfficesResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 2):
                for data in imported_data:
                    value = SchoolOffices(
                        data[0],
                        data[1]
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/uploaddb_schooloffices.html")


@login_required(login_url='login')
def upload_time(request):
    try:
        if request.method == 'POST':
            TimeResource()
            dataset = Dataset()
            new_students = request.FILES['myfile']

            imported_data = dataset.load(new_students.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_students)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row
            if(col == 3):
                for data in imported_data:
                    value = NewTime(
                        data[0],
                        data[1],
                        data[2]
                    )
                    value.save()
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
    except Exception:
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_time.html")

# admin


# teacher


@login_required(login_url='login')
def teacher_home_view(request, *args, **kwargs):
    global teacherNotif
    user = request.session.get('username')
    user_name = Faculty.objects.get(employee_id=user)
    fload = Facultyload.objects.filter(employee_id=user)
    offercode = OfferCode.objects.all()
    facultyload = []
    for object in offercode:
        for load in fload:
            if object.offer_code == load.offer_code_id:
                concatenatedStringofDays = ""
                for object1 in object.days:
                    concatenatedStringofDays += object1+" "
                facultyload.append(OfferCode(offer_code=object.offer_code, days=concatenatedStringofDays,
                                             start_time=object.start_time, end_time=object.end_time,
                                             room=object.room, subject_code=object.subject_code,
                                             sem_id=object.sem_id, academic_year=object.academic_year))
    count = 0
    numberOfNotif = NotificationFeedback.objects.filter(to_user=user)
    for check in numberOfNotif:
        if check.is_read == False:
            count = count + 1
    teacherNotif = count
    return render(request, "teacher/teacher_home.html",  {"object_list": facultyload, "teacherNotif": teacherNotif, "form": user_name})


@login_required(login_url='login')
def new(request, stud, id):
    global counselorNotif
    global studentNotif
    global teacherNotif
    time1 = ""
    time2 = ""
    subj = OfferCode.objects.get(offer_code=id)
    user = request.session.get('username')
    user_name = Faculty.objects.get(employee_id=user)
    studentReferred = AllStudent.objects.get(studnumber=stud)
    subject_referred = subj.subject_code
    form = TeachersReferralForm(instance=studentReferred, initial={
                                'subject_referred': subject_referred})
    qs = Facultyload.objects.filter(employee_id=user)
    degree = DegreeProgram.objects.get(
        program_id=studentReferred.degree_program_id)
    if request.method == "POST":
        reasons = request.POST['reasons']
        behavior = request.POST['behavior_problem']
        form = TeachersReferralForm(request.POST, instance=studentReferred, initial={
                                    'subject_referred': subject_referred})
        if form.is_valid():
            today = date.today()
            now = dt.datetime.now()
            ClassesofCounselor = []
            ClassesCounselor = []
            notAvailableSched = []
            sample = []
            tomorrow = today
            finder = 0
            all = Counselor.objects.all()
            # getting all the counselors
            for object in all:
                for object1 in object.program_designation:
                    if(degree.program_code == object1):
                        employeeId = object.employee_id
                        sample.append(Counselor(employee_id=object.employee_id, firstname=object.firstname,
                                                lastname=object.lastname, program_designation=object.program_designation))

            # counselor = Counselor.objects.get(
            #     program_designation=degree.program_code)
            OfferCodeCounselor = Facultyload.objects.filter(
                employee_id=employeeId)

            # counselor = []
            # OfferCodeCounselor = []
            # counselors = Counselor.objects.all()
            # for object in counselors:
            #     if object.program_designation is not None:
            #         for program in object.program_designation:
            #             if program == degree.program_code:
            #                 counselor.append(Counselor(employee_id=object.employee_id, firstname=object.firstname,
            #                                            lastname=object.lastname, program_designation=object.program_designation))

            # OfferCodeCounselors = Facultyload.objects.all()
            # for object in OfferCodeCounselors:
            #     for object1 in counselor:
            #         if object1.employee_id == object.employee_id:
            #             OfferCodeCounselor.append(Facultyload(id=object.id, employee_id=object.employee_id,
            #                                                   offer_code=object.offer_code))

            timeArray = []
            initialtime = 0

            newTime = str(initialtime)+':00:00'

            for x in range(24):
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                newTime = str(initialtime)+':30:00'
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                initialtime = initialtime + 1
                newTime = str(initialtime)+':00:00'

            while(finder == 0):
                tomorrow = tomorrow+timedelta(days=1)
                day_name = tomorrow.strftime("%a")
                if(day_name != "Sun" and day_name != "Sat"):
                    ScheduledReferralbyDay = TeachersReferral.objects.filter(
                        date=tomorrow)
                    OfferCodeCounselorCheck = bool(OfferCodeCounselor)
                    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)
                    if(ScheduledReferralbyDayCheck == True):
                        ScheduledReferralbyDayCheck = True
                    else:
                        ScheduledReferralbyDayCheck = False
                    if(OfferCodeCounselorCheck == True):
                        for object in OfferCodeCounselor:
                            Subject = OfferCode.objects.get(
                                offer_code=object.offer_code_id)
                            ClassesofCounselor.append(Subject)
                        for object in ClassesofCounselor:
                            for d in object.days:
                                if d == day_name:
                                    ClassesCounselor.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                                      start_time=object.start_time, end_time=object.end_time,
                                                                      room=object.room, subject_code=object.subject_code,
                                                                      sem_id=object.sem_id, academic_year=object.academic_year,))
                        ClassesCounselorCheck = bool(ClassesCounselor)
                    else:
                        ClassesCounselorCheck = False

                        # check the notAvailableSchedule

                    for a in ClassesCounselor:
                        print('oooooo', a.offer_code)
                    notAvailableSched = SetScheduleCounselor.objects.filter(
                        date=tomorrow)

                    notAvailableSchedCheck = bool(notAvailableSched)
                    print('notAvailableSchedCheck', notAvailableSchedCheck)
                    print('notAvailableSched', notAvailableSched)
                    for h in notAvailableSched:
                        print('hhhhhh', h.start_time, h.end_time)

                    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
                    end = datetime.strptime('17:30:00', '%H:%M:%S').time()

                    TimeTaken = 0
                    TimeTaken1 = 0
                    TimeTaken2 = 0
                    counter = 0

                    if (ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == False):
                        print('1')
                        startTime = datetime.strptime(
                            '8:00:00', '%H:%M:%S').time()
                        endTime = datetime.strptime(
                            '9:00:00', '%H:%M:%S').time()
                        time1 = startTime
                        time2 = endTime
                        finder = 1
                    elif(ClassesCounselorCheck == False and notAvailableSchedCheck == False and ScheduledReferralbyDayCheck == True):
                        print('2')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ScheduledReferralbyDay:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1

                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    TimeTaken = 0

                    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == False):
                        print('3')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ClassesCounselor:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    TimeTaken = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    TimeTaken = 0

                    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == True):
                        print('4')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in notAvailableSched:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    TimeTaken = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    TimeTaken = 0

                    elif(ClassesCounselorCheck == True and notAvailableSchedCheck == True and ScheduledReferralbyDayCheck == False):
                        print('5')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                print('suddd')
                                for obj in notAvailableSched:
                                    print('timailhan', obj.date)
                                    if(timeArray[x+1] <= obj.end_time and timeArray[x] >= obj.start_time):
                                        TimeTaken += 1
                                        print(
                                            "timailhan", obj.start_time, obj.end_time)

                                if(TimeTaken == 0):
                                    for object3 in ClassesCounselor:
                                        if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''
                        # notAvailableSched = []
                        # ClassesCounselor = []

                    elif(ClassesCounselorCheck == False and notAvailableSchedCheck == True and ScheduledReferralbyDayCheck == True):
                        print('6')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in notAvailableSched:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object3 in ScheduledReferralbyDay:
                                        if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''
                                # notAvailableSched = []
                                # ScheduledReferralbyDay = []

                    elif(ClassesCounselorCheck == True and notAvailableSchedCheck == False and ScheduledReferralbyDayCheck == True):
                        print('7')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ClassesCounselor:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object3 in ScheduledReferralbyDay:
                                        if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''
                        # ClassesCounselor = []
                        # ScheduledReferralbyDay = []

                    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == True):
                        print('8.s')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ScheduledReferralbyDay:
                                    print("1", timeArray[x])
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object3 in ClassesCounselor:
                                        print("2", timeArray[x])
                                        if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0):
                                        for object4 in notAvailableSched:
                                            print("3", timeArray[x])
                                            if(timeArray[x+1] <= object4.end_time and timeArray[x] >= object4.start_time):
                                                TimeTaken2 += 1
                                        if(TimeTaken2 == 0 and counter == 0):
                                            print('time1:', timeArray[x])
                                            time1 = timeArray[x]
                                            counter = 1
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                        elif(TimeTaken2 == 0 and counter == 1):
                                            print('time2:', timeArray[x])
                                            time2 = timeArray[x+1]
                                            finder = 1
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            break
                                        elif(TimeTaken2 != 0):
                                            time1 = ''
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            counter = 0
                                        else:
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1-0
                                            time1 = ''
                                    else:
                                        TimeTaken1 = 0
                                        counter = 0
                                        time1 = ''
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                                    # ScheduledReferralbyDay = []
                                    # ClassesCounselor = []
                                    # notAvailableSched = []

                    if(time1 != '' and time2 != ''):
                        print('9')
                        form.save()
                        behavior = form.cleaned_data['behavior_problem']
                        studentInfo = TeachersReferral(firstname=studentReferred.firstname,
                                                       lastname=studentReferred.lastname, studnumber=studentReferred.studnumber,
                                                       degree_program=degree.program_code, subject_referred=subject_referred,
                                                       reasons=reasons, counselor=employeeId, employeeid=user,
                                                       behavior_problem=behavior, start_time=time1, end_time=time2, date=tomorrow)
                        studentInfo.save()
                        form = TeachersReferralForm(instance=studentReferred, initial={
                                                    'subject_referred': subject_referred})
                        counselorNotif = counselorNotif + 1
                        # studentNotif = studentNotif + 1
                        create_notification(employeeId, user, 'manual_referral', extra_id=int(
                            stud), schedDay=tomorrow, schedStartTime=time1, schedEndTime=time2)
                        messages.info(
                            request, 'Successfully Referred the Student')
    return render(request, "teacher/refer_a_student.html", {"teacherNotif": teacherNotif, "form1": form, "form": user_name})


@login_required(login_url='login')
def teacher_view_students(request, id):
    global teacherNotif
    chuchu = list()
    chuchu = []
    studentslist = []
    finalstudentlist = []
    qs = Studentsload.objects.filter(offer_code=id)
    for std in qs:
        chuchu.append(AllStudent(std.studnumber_id))

    qs_student = AllStudent.objects.all()
    for stud in qs_student:
        for chu in chuchu:
            if stud.studnumber == chu.studnumber:
                studentslist.append(AllStudent(
                    stud.studnumber, stud.student_email))

    allstud = AllStudent.objects.all()
    for allstud in allstud:
        for stud in studentslist:
            if allstud.studnumber == stud.studnumber:
                finalstudentlist.append(AllStudent(
                    allstud.studnumber, allstud.firstname, allstud.lastname))
    user = request.session.get('username')
    user_name = Faculty.objects.get(employee_id=user)
    return render(request, "teacher/list_students.html", {"teacherNotif": teacherNotif, "id": id, "object_list": finalstudentlist, "form": user_name})


@login_required(login_url='login')
def teacher_view_referred_students(request, status):
    global teacherNotif
    user = request.session.get('username')
    if status == 'all' or status == '--':
        qs = TeachersReferral.objects.filter(employeeid=user)
    elif (status == 'pending'):
        qs = TeachersReferral.objects.filter(employeeid=user, status='pending')
    elif (status == 'done'):
        qs = TeachersReferral.objects.filter(employeeid=user, status='done')
    user_name = Faculty.objects.get(employee_id=user)
    filterform = FilterForm()
    if request.method == "POST":
        filterform = FilterForm(request.POST)
        if filterform.is_valid():
            filter_choice = filterform['filter_choice'].value()
            return redirect('teacher_view_referred_students', status=filter_choice)
    return render(request, "teacher/list_referred_students.html", {"status": status, "filterform": filterform, "teacherNotif": teacherNotif, "object_list": qs, "form": user_name})


@login_required(login_url='login')
def teacher_view_detail_referred_students(request, id):
    global teacherNotif
    user = request.session.get('username')
    detail = []
    qs = TeachersReferral.objects.filter(employeeid=user)
    for referedStud in qs:
        if(referedStud.id == id):
            detail.append(TeachersReferral(firstname=referedStud.firstname,
                                           lastname=referedStud.lastname, studnumber=referedStud.studnumber,
                                           degree_program=referedStud.degree_program, subject_referred=referedStud.subject_referred,
                                           reasons=referedStud.reasons, behavior_problem=referedStud.behavior_problem, feedback=referedStud.feedback))
    user_name = Faculty.objects.get(employee_id=user)
    return render(request, "teacher/info_referred_student.html", {"teacherNotif": teacherNotif, "object_list": detail, "form": user_name})


@login_required(login_url='login')
def teacher_coursecard(request, *args, **kwargs):
    user = request.session.get('username')
    qs = Facultyload.objects.get(employee_id=user)
    context = {"object_list": qs}
    return render(request, "teacher/about.html",)


@login_required
def notifications_teacher(request):
    user = request.session.get('username')
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = NotificationFeedback.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == NotificationFeedback.AUTOMATIC_REFERRAL:
            return render(request, "teacher/counselor.html", {})
        elif notification.notification_type == NotificationFeedback.MANUAL_REFERRAL:
            return render(request, "teacher/counselor.html", {})

    counselorNotif = NotificationFeedback.objects.filter(
        to_user=user, notification_type="feedback_teacher")
    user_name = Faculty.objects.get(employee_id=user)
    return render(request, 'teacher/notification.html', {"notifications": counselorNotif, "form": user_name})


@login_required(login_url='login')
def teacher_view_notif_detail(request, id):
    global teacherNotif
    user = request.session.get('username')
    detail = []
    notification = NotificationFeedback.objects.get(id=id)
    notification.is_read = True
    notification.save()
    # fback = CounselorFeedback.objects.get(id=notification.extra_id)
    students = TeachersReferral.objects.all()
    for referedStud in students:
        if notification.referral_id == referedStud.id:
            detail.append(TeachersReferral(firstname=referedStud.firstname,
                                           lastname=referedStud.lastname, studnumber=referedStud.studnumber,
                                           degree_program=referedStud.degree_program, subject_referred=referedStud.subject_referred,
                                           feedback=referedStud.feedback))
    user_name = Faculty.objects.get(employee_id=user)
    if teacherNotif != 0:
        teacherNotif = teacherNotif - 1
    return render(request, "teacher/detailNotif.html", {"objects": detail, "form": user_name})

# teacher

# counselor


@login_required(login_url='login')
def counselor_home_view(request, *args, **kwargs):
    global counselorNotif
    count = 0
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    numberOfNotif = Notification.objects.filter(to_user=user)
    for check in numberOfNotif:
        if check.is_read_counselor == False:
            count = count + 1
    counselorNotif = count
    return render(request, "counselor/counselor_home.html", {"counselorNotif": counselorNotif, "form": counselor_name})


@login_required(login_url='login')
def counselor_set_schedule(request, *args, **kwargs):
    global counselorNotif
    count = 0
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    numberOfNotif = Notification.objects.filter(to_user=user)

    for check in numberOfNotif:
        if check.is_read_counselor == False:
            count = count + 1
    # counselorNotif = counselor_setSchedule
    offer = SetScheduleCounselorForm(initial={
        'employee_id': user, 'choice': 'Not Available'})
    if request.method == "POST":
        offer = SetScheduleCounselorForm(request.POST, initial={
            'employee_id': user, 'choice': 'Not Available'})
        if offer.is_valid():
            pickedDateForm = offer['date'].value()
            start_timeForm = offer['start_time'].value()
            end_timeForm = offer['end_time'].value()

            date = datetime.strptime(pickedDateForm, "%Y-%m-%d").date()
            ClassesCounselor = []
            ClassesofCounselor = []
            ClassesCounselorCheck = False
            dateNotAvailable = date
            timeStart = datetime.strptime(
                start_timeForm + ':00', '%H:%M:%S').time()
            timeEnd = datetime.strptime(
                end_timeForm + ':00', '%H:%M:%S').time()
            day_name = date.strftime("%a")
            if timeEnd <= timeStart:
                # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                messages.info(request, 'sayop imong time')
            else:
                OfferCodeCounselor = Facultyload.objects.filter(
                    employee_id=user)
                OfferCodeCounselorChecker = bool(OfferCodeCounselor)

                if(OfferCodeCounselorChecker == True):
                    for object in OfferCodeCounselor:
                        Subject = OfferCode.objects.get(
                            offer_code=object.offer_code_id)
                        ClassesofCounselor.append(Subject)
                else:
                    OfferCodeCounselorChecker = False

                if(OfferCodeCounselorChecker == True):
                    for object in ClassesofCounselor:
                        for d in object.days:
                            if d == day_name:
                                ClassesCounselor.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                                  start_time=object.start_time, end_time=object.end_time, room=object.room,
                                                                  subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                                ClassesCounselorCheck = True
                else:
                    ClassesCounselorCheck = False
                notAvailableSched = SetScheduleCounselor.objects.filter(
                    employee_id=user, date=dateNotAvailable)
                notAvailableSchedChecker = bool(notAvailableSched)

                checker = 0
                checker1 = 0
                checker2 = 0
                ReferralNotAvailable = TeachersReferral.objects.filter(
                    date=dateNotAvailable).order_by('start_time')
                ReferralNotAvailableChecker = bool(ReferralNotAvailable)
                if(ReferralNotAvailableChecker == True and ClassesCounselorCheck == True and notAvailableSchedChecker == True):
                    print('1')
                    for object in ReferralNotAvailable:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                            print('1.1')
                    if(checker != 1):
                        for object1 in ClassesCounselor:
                            if(object1.start_time >= timeStart and object1.end_time <= timeEnd or object1.start_time <= timeStart and object1.end_time >= timeEnd):
                                checker1 = 1
                                print('1.2')
                        if(checker1 != 1):
                            for object2 in notAvailableSched:
                                if(object2.start_time >= timeStart and object2.end_time <= timeEnd or object2.start_time <= timeStart and object2.end_time >= timeEnd):
                                    checker2 = 1
                                    print('1.3')
                            if(checker2 != 1):
                                # -- save the id date and time in the database
                                offer.save()
                                newData = SetScheduleCounselor.objects.last()
                                newData.employee_id = user
                                newData.choice = 'Not Available'
                                newData.save()
                                # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                                messages.info(request, 'success')
                            else:
                                messages.info(
                                    request, 'ddi ka wede mag set kay naa kay klase ani')

                        else:
                            messages.info(
                                request, 'ddi ka wede mag set kay naa kay klase ani')

                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')

                    #ReferralNotAvailableChecker and ClassesCounselorCheck is true
                if(ReferralNotAvailableChecker == True and ClassesCounselorCheck == True and notAvailableSchedChecker == False):
                    print('2')
                    for object in ReferralNotAvailable:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                    if(checker != 1):
                        for object in ClassesCounselor:
                            if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.employee_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'success')
                        else:
                            messages.info(
                                request, 'ddi ka wede mag set kay naa kay klase ani')

                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')

                #ReferralNotAvailableChecker and notAvailableChecker is true
                if(ReferralNotAvailableChecker == True and ClassesCounselorCheck == False and notAvailableSchedChecker == True):
                    print('3')
                    for object in ReferralNotAvailable:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                    if(checker != 1):
                        for object in notAvailableSched:
                            if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeStart):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.employee_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'success')

                        else:
                            messages.info(
                                request, 'ddi ka wede mag set kay naa kay klase ani')

                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')

                #ClassesCounselorCheck and not AvailableChecker is true
                if(ReferralNotAvailableChecker == False and ClassesCounselorCheck == True and notAvailableSchedChecker == True):
                    print('4')
                    for object in ClassesCounselor:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                    if(checker != 1):
                        for object in notAvailableSched:
                            if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeStart):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.employee_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'success')
                        else:
                            messages.info(
                                request, 'ddi ka wede mag set kay naa kay klase ani')
                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')

                #ClassesCounselorCheck is True
                if(ReferralNotAvailableChecker == False and ClassesCounselorCheck == True and notAvailableSchedChecker == False):
                    print('5,s')
                    for object in ClassesCounselor:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                            print('a', object.start_time, object.end_time)
                            print('b', checker)
                    print('subject', ClassesCounselor)
                    print('timailan', checker)
                    if(checker != 1):
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.employee_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'success')
                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')
                #ReferralNotAvailableChecker is True
                if(ReferralNotAvailableChecker == True and ClassesCounselorCheck == False and notAvailableSchedChecker == False):
                    print('6')
                    for object in ClassesCounselor:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeEnd):
                            checker = 1
                    if(checker != 1):
                        # -- save the id date and time in the database
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.employee_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'success')
                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')
                #notAvailableSchedChecker is True
                if(ReferralNotAvailableChecker == True and ClassesCounselorCheck == False and notAvailableSchedChecker == False):
                    print('7')
                    for object in notAvailableSched:
                        if(object.start_time >= timeStart and object.end_time <= timeEnd or object.start_time <= timeStart and object.end_time >= timeStart):
                            checker = 1
                    if(checker != 1):
                        # -- save the id date and time in the database
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.employee_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'success')
                    else:
                        messages.info(
                            request, 'ddi ka wede mag set kay naa kay klase ani')
                # if all are false
                if(ReferralNotAvailableChecker == False and ClassesCounselorCheck == False and notAvailableSchedChecker == False):
                    print('8')
                    offer.save()
                    newData = SetScheduleCounselor.objects.last()
                    newData.employee_id = user
                    newData.choice = 'Not Available'
                    newData.save()
                    # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                    messages.info(request, 'success')

    return render(request, "counselor/counselor_set_schedule.html", {"offer": offer, "counselorNotif": counselorNotif, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_schedule(request, *args, **kwargs):
    global counselorNotif
    global count
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    count = 0
    user = request.session.get('username')
    today = date.today()
    now = dt.datetime.now()
    day_name = now.strftime("%a")
    ClassesofCounselor = []
    ClassesCounselor = []
    ClassesCounselorCheck = False
    ScheduledReferralbyDayCheck = False

    OfferCodeCounselor = Facultyload.objects.filter(employee_id=user)
    OfferCodeCounselorChecker = bool(OfferCodeCounselor)

    if(OfferCodeCounselorChecker == True):
        for object in OfferCodeCounselor:
            Subject = OfferCode.objects.get(offer_code=object.offer_code_id)
            ClassesofCounselor.append(Subject)
    else:
        OfferCodeCounselorChecker = False

    if(OfferCodeCounselorChecker == True):
        for object in ClassesofCounselor:
            for d in object.days:
                if d == day_name:
                    ClassesCounselor.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                      start_time=object.start_time, end_time=object.end_time, room=object.room,
                                                      subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                    ClassesCounselorCheck = True
    else:
        ClassesCounselorCheck = False

    ScheduledReferralbyDay = TeachersReferral.objects.filter(
        date=today).order_by('start_time')
    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)

    notAvailableSched = SetScheduleCounselor.objects.filter(
        date=today)
    notAvailableSchedCheck = bool(notAvailableSched)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    # one = []
    classForToday = []
    getFacultyload = Facultyload.objects.filter(employee_id=user)
    getOffercode = OfferCode.objects.all()
    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('7:00:00', '%H:%M:%S').time()
    end = datetime.strptime('18:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    getAllTime = NewTime.objects.all()
    alltime = []
    for obj in getAllTime:
        if obj.time1 >= start and obj.time2 < end and obj.time2 != check:
            alltime.append(NewTime(time1=obj.time1, time2=obj.time2))
    time = '00:00:00'
    if(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))
            for object4 in notAvailableSched:
                if(timeArray[x] == object4.start_time):
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object4.employee_id, date=object4.date, start_time=time,
                                                              end_time=object4.end_time, choice=choice))

    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in notAvailableSched:
                if(timeArray[x] >= object3.start_time and timeArray[x+1] <= object.end_time):
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object3.employee_id, date=object3.date, start_time=time,
                                                              end_time=object3.end_time, choice=choice))

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ClassesCounselor:
                if(timeArray[x] == object2.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object2.offer_code, days=object2.days,
                                                   start_time=time, end_time=object2.end_time, room=object2.room,
                                                   subject_code=object2.subject_code, sem_id=object2.sem_id, academic_year=object2.academic_year, choice=choice))

            for object3 in notAvailableSched:
                if(timeArray[x] >= object3.start_time and timeArray[x] <= object3.end_time):
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object3.employee_id, date=object3.date,
                                                              start_time=time, end_time=object3.end_time, choice=choice))

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))
    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(Time(time1=timeArray[x], time2=timeArray[x+1]))
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

        for object4 in notAvailableSched:
            if(timeArray[x] >= object3.start_time):
                choice = 'Not Available'
                classForToday.append(SetScheduleCounselor(employee_id=object4.employee_id, date=object4.date, start_time=time,
                                                          end_time=object4.end_time, choice=choice))

    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False):
        for x in getAllTime:
            if(x.time1 >= start and x.time2 <= end):
                # print('aaaa', x.time1)
                # print('bbb', x.time2)
                # one.append(Time(time1=x.time1, time2=x.time2))
                # print('one', one)
                time = x.time1

    counselor_name = Faculty.objects.get(employee_id=user)

    return render(request, "counselor/schedule.html", {"offer": offer, "counselorNotif": counselorNotif, "today": today, "day_name": day_name, "schedForToday": classForToday, "time": alltime, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_another_sched(request, *args, **kwargs):
    global counselorNotif
    global count
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    count = 0

    newDate = Calendar.objects.last()
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    today = newDate.pickedDate
    day_name = today.strftime("%a")
    ClassesofCounselor = []
    ClassesCounselor = []
    ClassesCounselorCheck = False
    ScheduledReferralbyDayCheck = False

    OfferCodeCounselor = Facultyload.objects.filter(employee_id=user)
    OfferCodeCounselorChecker = bool(OfferCodeCounselor)

    if(OfferCodeCounselorChecker == True):
        for object in OfferCodeCounselor:
            Subject = OfferCode.objects.get(offer_code=object.offer_code_id)
            ClassesofCounselor.append(Subject)
    else:
        OfferCodeCounselorChecker = False

    if(OfferCodeCounselorChecker == True):
        for object in ClassesofCounselor:
            for d in object.days:
                if d == day_name:
                    ClassesCounselor.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                      start_time=object.start_time, end_time=object.end_time, room=object.room,
                                                      subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                    ClassesCounselorCheck = True
    else:
        ClassesCounselorCheck = False

    ScheduledReferralbyDay = TeachersReferral.objects.filter(
        date=today).order_by('start_time')
    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)

    notAvailableSched = SetScheduleCounselor.objects.filter(
        date=today)
    notAvailableSchedCheck = bool(notAvailableSched)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    # one = []
    classForToday = []
    getFacultyload = Facultyload.objects.filter(employee_id=user)
    getOffercode = OfferCode.objects.all()

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('7:00:00', '%H:%M:%S').time()
    end = datetime.strptime('18:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    getAllTime = NewTime.objects.all()
    alltime = []
    for obj in getAllTime:
        if obj.time1 >= start and obj.time2 < end and obj.time2 != check:
            alltime.append(NewTime(time1=obj.time1, time2=obj.time2))

    time = '00:00:00'
    if(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == True):
        print('1')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))
            for object4 in notAvailableSched:
                if(timeArray[x] == object4.start_time):
                    print("sudtahay")
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object4.employee_id, date=object4.date, start_time=time,
                                                              end_time=object4.end_time, choice=choice))
        # for object5 in classForToday:
        #     print("sample", object5.start_time, object5.end_time)

    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == True):
        print('2')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in notAvailableSched:
                if(timeArray[x] == object3.start_time):
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object3.employee_id, date=object3.date, start_time=time,
                                                              end_time=object3.end_time, choice=choice))

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == False):
        for object1 in ScheduledReferralbyDay:
            print("sample", object1.firstname)
        print('3')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

        for object5 in classForToday:
            print("sample", object5.start_time, object5.end_time)

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == True):
        print('4')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ClassesCounselor:
                if(timeArray[x] == object2.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object2.offer_code, days=object2.days,
                                                   start_time=time, end_time=object2.end_time, room=object2.room,
                                                   subject_code=object2.subject_code, sem_id=object2.sem_id, academic_year=object2.academic_year, choice=choice))

            for object3 in notAvailableSched:
                if(timeArray[x] == object3.start_time):
                    choice = 'Not Available'
                    classForToday.append(SetScheduleCounselor(employee_id=object3.employee_id, date=object3.date,
                                                              start_time=time, end_time=object3.end_time, choice=choice))

    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == False):
        print('5')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))
    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == True and notAvailableSchedCheck == False):
        print('6')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))
    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False and notAvailableSchedCheck == True):
        print('7')
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesCounselor:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

        for object4 in notAvailableSched:
            if(timeArray[x] == object4.start_time):
                choice = 'Not Available'
                classForToday.append(SetScheduleCounselor(employee_id=object4.employee_id, date=object4.date, start_time=time,
                                                          end_time=object4.end_time, choice=choice))

    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] <= end):
                # one.append(timeArray[x])
                time = timeArray[x]

    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, "counselor/another_sched.html", {"offer": offer, "counselorNotif": counselorNotif, "today": today, "day_name": day_name, "schedForToday": classForToday, "time": alltime, "form": counselor_name})


@login_required(login_url='login')
def counselor_setSchedule(request, pk):
    global counselorNotif
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, "counselor/set_schedule.html", {"counselorNotif": counselorNotif, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_appointment(request, id):
    global counselorNotif
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    getappointmentToday = StudentSetSched.objects.get(id=id)
    return render(request, "counselor/appointment.html", {"counselorNotif": counselorNotif, "object": getappointmentToday, "form": counselor_name})


@login_required(login_url='login')
def counselor_detail_schedule_counseling(request, start, end, date):
    global counselorNotif
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    session = TeachersReferral.objects.get(
        start_time=start, end_time=end, date=date)
    return render(request, "counselor/schedule_information.html", {"counselorNotif": counselorNotif, "object": session, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_detail_referred_students(request, id):
    global counselorNotif
    user = request.session.get('username')
    notification = Notification.objects.get(id=id)
    notification.is_read_counselor = True
    notification.save()
    detail = []
    qs = TeachersReferral.objects.filter(counselor=user)
    type = "teacher"

    for referedStud in qs:
        if(referedStud.id == id):
            if referedStud.subject_referred is not None:
                type = "student"
                detail.append(TeachersReferral(id=id, firstname=referedStud.firstname,
                                               lastname=referedStud.lastname, studnumber=referedStud.studnumber,
                                               degree_program=referedStud.degree_program, subject_referred=referedStud.subject_referred,
                                               reasons=referedStud.reasons, behavior_problem=referedStud.behavior_problem, date=referedStud.date,
                                               start_time=referedStud.start_time, end_time=referedStud.end_time,
                                               status=referedStud.status, feedback=referedStud.feedback))
    user_name = Faculty.objects.get(employee_id=user)
    if counselorNotif != 0:
        counselorNotif = counselorNotif - 1
    return render(request, "counselor/students_referral_information.html", {"objects": detail, "type": type, "form": user_name})


@login_required(login_url='login')
def counselor_view_referred_students(request):
    global counselorNotif
    user = request.session.get('username')
    qs = TeachersReferral.objects.filter(counselor=user, status='done')
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, "counselor/referred_students.html", {"counselorNotif": counselorNotif, "objects": qs, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_pending_students(request, *args, **kwargs):
    global counselorNotif
    user = request.session.get('username')
    qs = TeachersReferral.objects.filter(counselor=user, status="pending")
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, "counselor/pending_students.html", {"counselorNotif": counselorNotif, "objects": qs, "form": counselor_name})


@login_required(login_url='login')
def counselor_feedback_student(request, id):
    global counselorNotif
    global teacherNotif
    global feedback_id
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(employee_id=user)
    info = Notification.objects.get(id=id)
    student = TeachersReferral.objects.get(id=id)
    preparedby = Faculty.objects.get(employee_id=student.counselor)
    form1 = CounselorFeedbackForm()
    if request.method == "POST":
        form1 = CounselorFeedbackForm(request.POST)
        if form1.is_valid():
            form1.save()
            feedback = form1['feedback'].value()
            feedback_id = feedback_id + 1
            messages.info(request, 'Successfully Feedback')
            t = TeachersReferral.objects.get(id=id)
            t.status = "done"
            t.save()
            t.feedback = feedback
            t.save()
            form1 = CounselorFeedbackForm()
            create_feedback(student.employeeid,
                            'feedback_teacher', user, feedback_id, id)

            return render(request, "counselor/feedback_student.html", {"counselorNotif": counselorNotif, "info": info,  "info2": preparedby,   "student": student, "object": form1, "form": counselor_name})

    return render(request, "counselor/feedback_student.html", {"counselorNotif": counselorNotif, "info": info, "info2": preparedby, "student": student, "object": form1, "form": counselor_name})


@login_required(login_url='login')
def counselor_feedback(request, id):
    global counselorNotif
    global teacherNotif
    global feedback_id
    user = request.session.get('username')
    check = TeachersReferral.objects.all()
    faculty = Faculty.objects.all()
    flagteacher = 0
    for exist in check:
        for teacher in faculty:
            if exist.employeeid == teacher.employee_id:
                flagteacher = 1

    counselor_name = Faculty.objects.get(employee_id=user)
    info = Notification.objects.get(id=id)
    student = TeachersReferral.objects.get(id=id)
    referredby = Faculty.objects.get(employee_id=student.employeeid)
    preparedby = Faculty.objects.get(employee_id=student.counselor)
    form1 = CounselorFeedbackForm()
    if request.method == "POST":
        form1 = CounselorFeedbackForm(request.POST)
        if form1.is_valid():
            form1.save()
            feedback = form1['feedback'].value()
            feedback_id = feedback_id + 1
            if flagteacher == 1:
                create_feedback(student.employeeid,
                                'feedback_teacher', user, feedback_id, id)

            messages.info(request, 'Successfully Feedback')
            # teacherNotif = teacherNotif + 1
            t = TeachersReferral.objects.get(id=id)
            t.status = "done"
            t.save()
            t.feedback = feedback
            t.save()
            form1 = CounselorFeedbackForm()
            return render(request, "counselor/feedback.html", {"counselorNotif": counselorNotif,
                                                               "info": info,  "info1": referredby, "info2": preparedby,   "student": student, "object": form1,
                                                               "form": counselor_name})
    return render(request, "counselor/feedback.html", {"counselorNotif": counselorNotif, "info": info,
                                                       "info1": referredby, "info2": preparedby, "student": student, "object": form1, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_feedback(request):
    global counselorNotif
    user = request.session.get('username')
    student = TeachersReferral.objects.filter(counselor=user, status='done')
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, 'counselor/view_feedback.html', {"offer": offer, "counselorNotif": counselorNotif, "student": student, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_another_feedback(request):
    global counselorNotif
    user = request.session.get('username')
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    newDate = Calendar.objects.last()

    counselor_name = Faculty.objects.get(employee_id=user)
    today = newDate.pickedDate
    student = TeachersReferral.objects.filter(
        counselor=user, status='done', date=today)
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, 'counselor/view_another_feedback.html', {"offer": offer, "counselorNotif": counselorNotif, "student": student, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_detail_feedback(request, id):
    global counselorNotif
    user = request.session.get('username')
    student = TeachersReferral.objects.filter(id=id)
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, 'counselor/detail_feedback.html', {"counselorNotif": counselorNotif, "student": student, "form": counselor_name})


@login_required
def notifications(request):
    user = request.session.get('username')
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = Notification.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == Notification.AUTOMATIC_REFERRAL:
            return render(request, "counselor/counselor.html", {})
        elif notification.notification_type == Notification.MANUAL_REFERRAL:
            return render(request, "counselor/counselor.html", {})

    counselorNotif = Notification.objects.filter(to_user=user)
    counselor_name = Faculty.objects.get(employee_id=user)
    return render(request, 'counselor/notification.html', {"notifications": counselorNotif, "form": counselor_name})
# counselor

# student


@login_required(login_url='login')
def student_add_info(request, *args, **kwargs):
    user = request.session.get('username')
    student_name = AllStudent.objects.get(studnumber=user)
    infoForm = StudentInfoForm(instance=student_name)
    if request.method == "POST":
        infoForm = StudentInfoForm(request.POST, instance=student_name)
        mother_firstname = request.POST['mother_firstname']
        mother_lastname = request.POST['mother_lastname']
        father_firstname = request.POST['father_firstname']
        father_lastname = request.POST['father_lastname']
        guardian_firstname = request.POST['guardian_firstname']
        guardian_lastname = request.POST['guardian_lastname']
        student_contact_number = request.POST['student_contact_number']
        mother_contact_number = request.POST['mother_contact_number']
        father_contact_number = request.POST['father_contact_number']
        guardian_contact_number = request.POST['guardian_contact_number']
        if infoForm.is_valid():
            infoForm.save()
            info = StudentInfo.objects.get(studnumber=user)
            info.mother_firstname = mother_firstname
            info.mother_lastname = mother_lastname
            info.father_firstname = father_firstname
            info.father_lastname = father_lastname
            info.guardian_firstname = guardian_firstname
            info.guardian_lastname = guardian_lastname
            info.student_contact_number = student_contact_number
            info.mother_contact_number = mother_contact_number
            info.father_contact_number = father_contact_number
            info.guardian_contact_number = guardian_contact_number
            info.status = "done"
            info.save()
            return redirect('student_home_view')
    return render(request, "student/add_info.html", {"form": student_name, "info": infoForm})


@login_required(login_url='login')
def student_home_view(request, *args, **kwargs):
    global studentNotif
    count = 0
    user = request.session.get('username')
    numberOfNotif = Notification.objects.filter(extra_id=user)
    for check in numberOfNotif:
        if check.is_read_student == False:
            count = count + 1
    studentNotif = count
    student_name = AllStudent.objects.get(studnumber=user)
    return render(request, "student/student_home.html", {"studentNotif": studentNotif, "form": student_name})


@login_required(login_url='login')
def student_edit_profile_view(request, *args, **kwargs):
    user = request.session.get('username')
    student = StudentInfo.objects.get(studnumber=user)
    infoForm = StudentInfoForm(instance=student)
    if request.method == "POST":
        infoForm = StudentInfoForm(request.POST, instance=student)
        mother_firstname = request.POST['mother_firstname']
        mother_lastname = request.POST['mother_lastname']
        father_firstname = request.POST['father_firstname']
        father_lastname = request.POST['father_lastname']
        guardian_firstname = request.POST['guardian_firstname']
        guardian_lastname = request.POST['guardian_lastname']
        student_contact_number = request.POST['student_contact_number']
        mother_contact_number = request.POST['mother_contact_number']
        father_contact_number = request.POST['father_contact_number']
        guardian_contact_number = request.POST['guardian_contact_number']
        if infoForm.is_valid():
            infoForm.save()
            info = StudentInfo.objects.get(studnumber=user)
            info.mother_firstname = mother_firstname
            info.mother_lastname = mother_lastname
            info.father_firstname = father_firstname
            info.father_lastname = father_lastname
            info.guardian_firstname = guardian_firstname
            info.guardian_lastname = guardian_lastname
            info.student_contact_number = student_contact_number
            info.mother_contact_number = mother_contact_number
            info.father_contact_number = father_contact_number
            info.guardian_contact_number = guardian_contact_number
            info.save()
            messages.info(request, 'Successfully edit')
    return render(request, "student/student_edit_profile_view.html", {"form": student, "info": infoForm})


@login_required(login_url='login')
def student_schedule(request, *args, **kwargs):
    global counselorNotif
    global studentNotif
    user = request.session.get('username')
    student_name = AllStudent.objects.get(studnumber=user)
    schedForm = StudentSetSchedForm(instance=student_name)
    degree = DegreeProgram.objects.get(
        program_id=student_name.degree_program_id)
    if request.method == "POST":
        reasons = request.POST['reasons']
        schedForm = StudentSetSchedForm(request.POST, instance=student_name)
        if schedForm.is_valid():
            today = date.today()
            now = dt.datetime.now()
            ClassesofCounselor = []
            ClassesCounselor = []
            tomorrow = today
            finder = 0

            # counselor = Counselor.objects.get(
            #     program_designation=degree.program_code)
            get_counselor = Counselor.objects.all()
            for object in get_counselor:
                for object1 in object.program_designation:
                    if(degree.program_code in object1):
                        counselor = object

            OfferCodeCounselor = Facultyload.objects.filter(
                employee_id=counselor.employee_id)
            timeArray = []
            initialtime = 0

            newTime = str(initialtime)+':00:00'

            for x in range(24):
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                newTime = str(initialtime)+':30:00'
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                initialtime = initialtime + 1
                newTime = str(initialtime)+':00:00'

            while(finder == 0):
                tomorrow = tomorrow+timedelta(days=1)
                day_name = tomorrow.strftime("%a")
                if(day_name != "Sun" and day_name != "Sat"):
                    ScheduledReferralbyDay = TeachersReferral.objects.filter(
                        date=tomorrow)
                    OfferCodeCounselorCheck = bool(OfferCodeCounselor)
                    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)
                    if(ScheduledReferralbyDayCheck == True):
                        ScheduledReferralbyDayCheck = True
                    else:
                        ScheduledReferralbyDayCheck = False
                    if(OfferCodeCounselorCheck == True):
                        for object in OfferCodeCounselor:
                            Subject = OfferCode.objects.get(
                                offer_code=object.offer_code_id)
                            ClassesofCounselor.append(Subject)
                        for object in ClassesofCounselor:
                            for d in object.days:
                                if d == day_name:
                                    ClassesCounselor.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                                      start_time=object.start_time, end_time=object.end_time, room=object.room,
                                                                      subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                        ClassesCounselorCheck = bool(ClassesCounselor)
                    else:
                        ClassesCounselorCheck = False

                    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
                    end = datetime.strptime('17:00:00', '%H:%M:%S').time()

                    TimeTaken = 0
                    TimeTaken1 = 0
                    counter = 0

                    if (ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == False):
                        startTime = datetime.strptime(
                            '8:00:00', '%H:%M:%S').time()
                        endTime = datetime.strptime(
                            '9:00:00', '%H:%M:%S').time()
                        time1 = startTime
                        time2 = endTime
                        finder = 1
                    elif(ClassesCounselorCheck == False and ScheduledReferralbyDayCheck == True):
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ScheduledReferralbyDay:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1

                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    TimeTaken = 0

                    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == False):
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ClassesCounselor:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    TimeTaken = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    TimeTaken = 0
                    elif(ClassesCounselorCheck == True and ScheduledReferralbyDayCheck == True):
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object2 in ScheduledReferralbyDay:
                                    if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object3 in ClassesCounselor:
                                        if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''
                    ScheduledReferralbyDay = []
                    ClassesCounselor = []
            if(time1 != '' and time2 != ''):
                schedForm.save()
                # qs = Counselor.objects.get(
                #     program_designation=degree.program_code)
                sched = StudentSetSched(studnumber=student_name.studnumber,
                                        firstname=student_name.firstname, lastname=student_name.lastname,
                                        degree_program=degree.program_code,
                                        reasons=reasons, counselor=counselor.employee_id,
                                        start_time=time1, end_time=time2, date=tomorrow)
                sched.save()
                studentInfo = TeachersReferral(firstname=student_name.firstname,
                                               lastname=student_name.lastname, studnumber=student_name.studnumber,
                                               degree_program=degree.program_code,
                                               reasons=reasons, counselor=counselor.employee_id,
                                               employeeid=user, start_time=time1, end_time=time2, date=tomorrow)
                studentInfo.save()
                schedForm = StudentSetSchedForm(instance=student_name)
                context = {"schedform": schedForm, "form": student_name}
                counselorNotif = counselorNotif + 1
                studentNotif = studentNotif + 1
                create_notification(counselor.employee_id, user, 'appointment', extra_id=int(
                    student_name.studnumber), schedDay=tomorrow, schedStartTime=time1, schedEndTime=time2)
                messages.info(request, 'Successfully Set Schedule')

    return render(request, "student/schedule.html", {"studentNotif": studentNotif, "schedform": schedForm, "form": student_name})


@login_required
def view_schedule_student(request, *args, **kwargs):
    global count1
    count1 = 0
    global studentNotif
    user = request.session.get('username')
    today = date.today()
    now = dt.datetime.now()
    day_name = now.strftime("%a")

    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    # newDate = Calendar.objects.last()

    # user = request.session.get('username')
    # counselor_name = Faculty.objects.get(employee_id=user)
    # today = newDate.pickedDate
    # day_name = today.strftime("%a")
    ClassesofStudent = []
    ClassesStudent = []
    ClassesStudentCheck = False
    ScheduledReferralbyDayCheck = False

    OfferCodeStudent = Studentsload.objects.filter(studnumber=user)
    OfferCodeStudentChecker = bool(OfferCodeStudent)

    if(OfferCodeStudentChecker == True):
        for object in OfferCodeStudent:
            Subject = OfferCode.objects.get(offer_code=object.offer_code_id)
            ClassesofStudent.append(Subject)
    else:
        OfferCodeStudentChecker = False

    if(OfferCodeStudentChecker == True):
        for object in ClassesofStudent:
            for d in object.days:
                if d == day_name:
                    ClassesStudent.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                    start_time=object.start_time, end_time=object.start_time, room=object.room,
                                                    subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                    ClassesStudentCheck = True
    else:
        ClassesStudentCheck = False

    ScheduledReferralbyDay = TeachersReferral.objects.filter(
        date=today).order_by('start_time')
    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    one = []
    two = []
    three = []
    getsched = []
    classForToday = []

    getStudentsload = Studentsload.objects.filter(studnumber=user)
    getOffercode = OfferCode.objects.all()

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
    end = datetime.strptime('17:00:00', '%H:%M:%S').time()

    if(ClassesStudentCheck == True and ScheduledReferralbyDayCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == False and ScheduledReferralbyDayCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == True and ScheduledReferralbyDayCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == False and ScheduledReferralbyDayCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]

    student_name = AllStudent.objects.get(studnumber=user)
    return render(request, "student/viewschedule.html", {"studentNotif": studentNotif, "offer": offer, "today": today, "day_name": day_name, "schedForToday": classForToday, "time": one, "form": student_name})


@login_required(login_url='login')
def student_view_another_sched(request):
    global count1
    # count1 += num
    global studentNotif
    user = request.session.get('username')
    # today = date.today() + timedelta(days=count1)
    # now = dt.datetime.now() + timedelta(days=count1)
    # day_name = now.strftime("%a")

    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()
    newDate = Calendar.objects.last()
    today = newDate.pickedDate
    day_name = today.strftime("%a")

    ClassesofStudent = []
    ClassesStudent = []
    ClassesStudentCheck = False
    ScheduledReferralbyDayCheck = False

    OfferCodeStudent = Studentsload.objects.filter(studnumber=user)
    OfferCodeStudentChecker = bool(OfferCodeStudent)

    if(OfferCodeStudentChecker == True):
        for object in OfferCodeStudent:
            Subject = OfferCode.objects.get(offer_code=object.offer_code_id)
            ClassesofStudent.append(Subject)
    else:
        OfferCodeStudentChecker = False

    if(OfferCodeStudentChecker == True):
        for object in ClassesofStudent:
            for d in object.days:
                if d == day_name:
                    ClassesStudent.append(OfferCode(offer_code=object.offer_code, days=object.days,
                                                    start_time=object.start_time, end_time=object.start_time, room=object.room,
                                                    subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                    ClassesStudentCheck = True
    else:
        ClassesStudentCheck = False

    ScheduledReferralbyDay = TeachersReferral.objects.filter(
        date=today).order_by('start_time')
    ScheduledReferralbyDayCheck = bool(ScheduledReferralbyDay)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    one = []
    two = []
    three = []
    getsched = []
    classForToday = []

    getStudentsload = Studentsload.objects.filter(studnumber=user)
    getOffercode = OfferCode.objects.all()

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
    end = datetime.strptime('17:00:00', '%H:%M:%S').time()

    if(ClassesStudentCheck == True and ScheduledReferralbyDayCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == False and ScheduledReferralbyDayCheck == True):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == True and ScheduledReferralbyDayCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]
            for object2 in ScheduledReferralbyDay:
                if(timeArray[x] == object2.start_time):
                    choice = 'Counseling'
                    classForToday.append(TeachersReferral(firstname=object2.firstname,
                                                          lastname=object2.lastname, studnumber=object2.studnumber,
                                                          degree_program=object2.degree_program, subject_referred=object2.subject_referred,
                                                          reasons=object2.reasons, behavior_problem=object2.behavior_problem, date=object2.date,
                                                          start_time=time, end_time=object2.end_time, choice=choice))

            for object3 in ClassesStudent:
                if(timeArray[x] == object3.start_time):
                    choice = 'Class'
                    classForToday.append(OfferCode(offer_code=object3.offer_code, days=object3.days,
                                                   start_time=time, end_time=object3.end_time, room=object3.room,
                                                   subject_code=object3.subject_code, sem_id=object3.sem_id, academic_year=object3.academic_year, choice=choice))

    elif(ClassesStudentCheck == False and ScheduledReferralbyDayCheck == False):
        for x in range(len(timeArray)):
            if(timeArray[x] >= start and timeArray[x] < end):
                one.append(timeArray[x])
                time = timeArray[x]

    student_name = AllStudent.objects.get(studnumber=user)
    return render(request, "student/another_sched.html", {"studentNotif": studentNotif, "offer": offer, "today": today, "day_name": day_name, "schedForToday": classForToday, "time": one, "form": student_name})


@login_required
def view_appointment_students(request, start, end, date):
    global studentNotif
    user = request.session.get('username')
    student_name = AllStudent.objects.get(studnumber=user)
    getStudentsSched = TeachersReferral.objects.get(
        start_time=start, end_time=end, date=date)
    return render(request, "student/viewappointment.html", {"studentNotif": studentNotif, "object": getStudentsSched, "form": student_name})


@login_required
def notifications_student(request):
    user = request.session.get('username')
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = Notification.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == Notification.AUTOMATIC_REFERRAL:
            return render(request, "student/student_home.html", {})
        elif notification.notification_type == Notification.MANUAL_REFERRAL:
            return render(request, "student/student_home", {})
    notif = Notification.objects.filter(extra_id=user)
    student_name = AllStudent.objects.get(studnumber=user)
    return render(request, 'student/notification.html', {"notifications": notif, "form": student_name})


@login_required
def student_notif_detail(request, id):
    global studentNotif
    user = request.session.get('username')
    notification = Notification.objects.get(id=id)
    notification.is_read_student = True
    notification.save()
    detail = []
    student = TeachersReferral.objects.all()
    for referedStud in student:
        if (referedStud.id == id):
            detail.append(TeachersReferral(firstname=referedStud.firstname,
                                           lastname=referedStud.lastname, studnumber=referedStud.studnumber,
                                           degree_program=referedStud.degree_program, subject_referred=referedStud.subject_referred,
                                           reasons=referedStud.reasons, behavior_problem=referedStud.behavior_problem,
                                           date=referedStud.date, start_time=referedStud.start_time, end_time=referedStud.end_time))
    if studentNotif != 0:
        studentNotif = studentNotif - 1
    student_name = AllStudent.objects.get(studnumber=user)
    return render(request, 'student/notif_detail.html', {"studentNotif": studentNotif, "object_list": detail, "form": student_name})


@login_required
def student_history(request):
    global studentNotif
    user = request.session.get('username')
    student_name = AllStudent.objects.get(studnumber=user)
    student_record = TeachersReferral.objects.filter(
        studnumber=user, status='done')
    counselor = Counselor.objects.all()

    return render(request, "student/view_history.html", {"studentNotif": studentNotif, "object": student_record, "counselor": counselor, "form": student_name})
